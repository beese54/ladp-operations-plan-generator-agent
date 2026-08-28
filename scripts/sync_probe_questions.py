"""Sync the RAG probe question bank between Excel (SME editing) and CSV (source of truth).

Why both formats
----------------
SMEs fill in ground-truth answers far more comfortably in Excel, and it removes a
real failure mode: the first SME pass corrupted a row because an answer contained
a comma, which shifted it into the neighbouring CSV column.

But .xlsx is binary — `git diff` shows nothing and merge conflicts can't be
resolved by hand. So the CSV stays the committed source of truth (reviewable in a
PR, loadable by evaluation/datasets.py) and Excel is only an editing surface.

Workflow
--------
    # 1. Generate an Excel workbook for the SME to fill in
    PYTHONPATH=. python scripts/sync_probe_questions.py to-excel

    # 2. SME edits the .xlsx, adding ground_truth answers

    # 3. Pull their answers back into the canonical CSV
    PYTHONPATH=. python scripts/sync_probe_questions.py to-csv

    # Check what would change without writing anything
    PYTHONPATH=. python scripts/sync_probe_questions.py to-csv --dry-run

Both directions validate the data and report problems rather than silently
writing a malformed bank.
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd

CSV_PATH = Path("data/eval_datasets/rag_probe_questions.csv")
XLSX_PATH = Path("data/eval_datasets/rag_probe_questions.xlsx")
SHEET = "questions"

REQUIRED_COLUMNS = [
    "id", "persona", "difficulty", "category", "question",
    "should_answer", "expected_behaviour", "rag_gap_hypothesis",
    "sme_verified", "ground_truth",
]

VALID_PERSONAS = {"OPS_PLANNER", "FIELD_CREW", "GENERAL", "OUT_OF_SCOPE"}
VALID_SHOULD_ANSWER = {"YES", "NO"}
VALID_SME = {"YES", "NO"}

# Column widths that make the sheet actually usable for an SME.
COLUMN_WIDTHS = {
    "id": 6, "persona": 14, "difficulty": 11, "category": 18,
    "question": 60, "should_answer": 13, "expected_behaviour": 45,
    "rag_gap_hypothesis": 30, "sme_verified": 13, "ground_truth": 80,
}


def _load_csv() -> pd.DataFrame:
    if not CSV_PATH.exists():
        sys.exit(f"CSV not found: {CSV_PATH}")
    # keep_default_na=False so blank ground_truth reads as "" not NaN — otherwise
    # every empty answer round-trips into the literal string "nan".
    return pd.read_csv(CSV_PATH, keep_default_na=False, dtype=str)


def _load_excel() -> pd.DataFrame:
    if not XLSX_PATH.exists():
        sys.exit(f"Excel not found: {XLSX_PATH}\n"
                 f"Run 'to-excel' first to generate it.")
    df = pd.read_excel(XLSX_PATH, sheet_name=SHEET, dtype=str)
    return df.fillna("")


def validate(df: pd.DataFrame) -> list[str]:
    """Return a list of human-readable problems. Empty list means clean."""
    problems: list[str] = []

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing_cols:
        problems.append(f"missing column(s): {', '.join(missing_cols)}")
        return problems  # further checks would all fail spuriously

    if df["id"].duplicated().any():
        dupes = sorted(df.loc[df["id"].duplicated(), "id"])
        problems.append(f"duplicate id(s): {', '.join(dupes)}")

    for _, row in df.iterrows():
        rid = row["id"] or "(blank id)"

        if not str(row["question"]).strip():
            problems.append(f"{rid}: question is empty")

        persona = str(row["persona"]).strip().upper()
        if persona not in VALID_PERSONAS:
            problems.append(f"{rid}: persona {persona!r} not in {sorted(VALID_PERSONAS)}")

        should = str(row["should_answer"]).strip().upper()
        if should not in VALID_SHOULD_ANSWER:
            problems.append(f"{rid}: should_answer {should!r} must be YES or NO")

        sme = str(row["sme_verified"]).strip().upper()
        if sme not in VALID_SME:
            problems.append(f"{rid}: sme_verified {sme!r} must be YES or NO")

        # The pair that actually matters for scoring: a row claiming SME
        # verification with no answer would be silently skipped by the
        # correctness scorer, so surface it loudly.
        gt = str(row["ground_truth"]).strip()
        if sme == "YES" and not gt:
            problems.append(f"{rid}: sme_verified=YES but ground_truth is empty")
        if sme == "NO" and gt:
            problems.append(f"{rid}: has ground_truth but sme_verified=NO "
                            f"(set it to YES so the answer gets scored)")

        # Out-of-scope rows are graded on refusal, not on content.
        # "DECLINE" is the special ground_truth value for these — it means
        # the expected behaviour is a polite refusal, not a substantive answer.
        if should == "NO" and gt and gt.upper() != "DECLINE":
            problems.append(f"{rid}: should_answer=NO should not have a ground_truth answer "
                            f"(use DECLINE for refusal rows)")

    return problems


def summarise(df: pd.DataFrame) -> None:
    total = len(df)
    verified = (df["sme_verified"].str.strip().str.upper() == "YES").sum()
    print(f"  rows: {total}")
    print(f"  SME-verified with ground truth: {verified}")
    print(f"  awaiting SME answer: {total - verified}")
    by_persona = df["persona"].str.strip().str.upper().value_counts()
    for persona, n in by_persona.items():
        sub = df[df["persona"].str.strip().str.upper() == persona]
        v = (sub["sme_verified"].str.strip().str.upper() == "YES").sum()
        print(f"    {persona:<14} {n:>3} rows, {v:>2} verified")


def to_excel(dry_run: bool) -> None:
    df = _load_csv()
    problems = validate(df)
    if problems:
        print("Validation problems in the CSV:")
        for p in problems:
            print(f"  - {p}")
        print()

    print(f"Source: {CSV_PATH}")
    summarise(df)

    if dry_run:
        print(f"\n--dry-run: would write {XLSX_PATH}")
        return

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=SHEET, index=False)
        ws = writer.sheets[SHEET]

        # Freeze the header and the id/question columns so an SME scrolling right
        # to the ground_truth column can still see which question they're answering.
        ws.freeze_panes = "B2"

        from openpyxl.styles import Alignment, Font, PatternFill

        header_fill = PatternFill("solid", fgColor="1E293B")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        for idx, col in enumerate(df.columns, start=1):
            letter = ws.cell(row=1, column=idx).column_letter
            ws.column_dimensions[letter].width = COLUMN_WIDTHS.get(col, 20)
            # Wrap the two long free-text columns so answers are readable.
            if col in ("question", "expected_behaviour", "ground_truth"):
                for r in range(2, len(df) + 2):
                    ws.cell(row=r, column=idx).alignment = Alignment(
                        wrap_text=True, vertical="top")

    print(f"\nWrote {XLSX_PATH}")
    print("Send this to the SME. They only need to fill 'ground_truth' and set "
          "'sme_verified' to YES on rows they answer.")


def to_csv(dry_run: bool) -> None:
    df = _load_excel()

    # Excel silently drops leading/trailing spaces inconsistently; normalise so
    # the committed CSV is stable and diffs stay meaningful.
    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()

    problems = validate(df)
    if problems:
        print("Validation problems in the Excel file:")
        for p in problems:
            print(f"  - {p}")
        print("\nFix these in the .xlsx and re-run. Nothing was written.")
        sys.exit(1)

    df = df[REQUIRED_COLUMNS]  # canonical column order

    print(f"Source: {XLSX_PATH}")
    summarise(df)

    if CSV_PATH.exists():
        before = _load_csv()
        if len(before) == len(df):
            changed = [
                str(df.iloc[i]["id"])
                for i in range(len(df))
                if not df.iloc[i].equals(before.iloc[i])
            ]
            print(f"\n  rows changed: {len(changed)}"
                  + (f" ({', '.join(changed)})" if changed else ""))
        else:
            print(f"\n  row count changed: {len(before)} -> {len(df)}")

    if dry_run:
        print(f"\n--dry-run: would write {CSV_PATH}")
        return

    df.to_csv(CSV_PATH, index=False, lineterminator="\n")
    print(f"\nWrote {CSV_PATH}")


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("direction", choices=["to-excel", "to-csv"],
                    help="to-excel: CSV -> xlsx for SME editing. "
                         "to-csv: pull SME answers back into the canonical CSV.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Validate and report, but write nothing")
    args = ap.parse_args()

    if args.direction == "to-excel":
        to_excel(args.dry_run)
    else:
        to_csv(args.dry_run)


if __name__ == "__main__":
    main()
